# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active  # type: ignore

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]
# Remover uma planilha
workbook.remove(workbook['Sheet'])


# criando os cabecalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome-idade-nota
    ['Joao', 14, 5.5],
    ['Maria', 14, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10],
]


# =======================================================
# Adicionando os dados dos alunos
# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)  # type: ignore
# =======================================================

for student in students:
    worksheet.append(student)

# =======================================================
# Adicionando os dados dos alunos
# for i, student in enumerate(students, start=2):
#     worksheet.cell(i, 1, student[0])
#     worksheet.cell(i, 2, student[1])
#     worksheet.cell(i, 3, student[2])


# =======================================================
# for i in range(2, 10):
#     for j in range(1, 4):
#         print('linha', i, 'coluna', j)
# =======================================================


workbook.save(WORKBOOK_PATH)  # type: ignore

# python -m pip install types-openpyxl
